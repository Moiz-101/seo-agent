"""Full evidence tables as XLSX files - the PDF stays a concise, client-
readable summary; anyone who wants the complete row-by-row data (every
broken link occurrence, every crawled page, etc.) gets it here instead of
a 25-page PDF appendix.
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import DATA_DIR
from seo_agent.research.site_crawler import compute_page_issues

XLSX_DIR = os.path.join(DATA_DIR, "audit_xlsx")
os.makedirs(XLSX_DIR, exist_ok=True)

HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    for i, header in enumerate(headers, start=1):
        width = max(12, min(60, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _safe_name(url: str) -> str:
    return "".join(c if c.isalnum() else "-" for c in url).strip("-")


def build_broken_links_xlsx(site_data: dict) -> str:
    report = site_data.get("broken_links_report", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "Confirmed Broken"
    headers = ["Source Page", "Destination", "Status", "Type", "Anchor Text", "Suggested Redirect", "Recommended Fix"]
    rows = [
        [
            b["source_page"], b["destination"], str(b["status"]), b["type"], b["anchor_text"],
            b.get("suggested_redirect") or "",
            "301-redirect to the suggested URL" if b.get("suggested_redirect") else ("Update or remove this link" if b["type"] == "external" else "Fix the link or set up a 301 redirect"),
        ]
        for b in report.get("confirmed_broken", [])
    ]
    _write_sheet(ws, headers, rows)

    ws2 = wb.create_sheet("Unverified")
    headers2 = ["Source Page", "Destination", "Status/Reason", "Failure Reason", "Type", "Anchor Text"]
    rows2 = [
        [u["source_page"], u["destination"], str(u["status"]), u.get("failure_reason") or "", u["type"], u["anchor_text"]]
        for u in report.get("unverified", [])
    ]
    _write_sheet(ws2, headers2, rows2)

    path = os.path.join(XLSX_DIR, f"broken-links-{_safe_name(site_data['url'])}.xlsx")
    wb.save(path)
    return path


def build_image_alt_xlsx(site_data: dict) -> str:
    findings = site_data.get("image_alt_findings", [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Alt Text"
    headers = ["Image URL", "Occurrence Count", "Sitewide Reused", "Classification", "Priority", "Recommended Alt", "AI-Generated", "Sample Page"]
    rows = [
        [
            f["image_url"], f["occurrence_count"], "Yes" if f.get("sitewide_reused") else "No",
            f["classification"], f["priority"], f.get("recommended_alt") or "",
            "Yes" if f.get("recommended_alt_ai_generated") else "No", f.get("page_url", ""),
        ]
        for f in findings
    ]
    _write_sheet(ws, headers, rows)
    path = os.path.join(XLSX_DIR, f"image-alt-audit-{_safe_name(site_data['url'])}.xlsx")
    wb.save(path)
    return path


def build_page_audit_xlsx(site_data: dict) -> str:
    pages = site_data["crawl"]["pages"]
    broken_source_pages = {b["source_page"] for b in site_data.get("broken_links_report", {}).get("confirmed_broken", [])}
    protocol_dup_urls = {d["http_url"] for d in site_data.get("protocol_issues", {}).get("protocol_duplicates", [])}

    wb = Workbook()
    ws = wb.active
    ws.title = "Page Audit"
    headers = [
        "URL", "Status", "Final URL", "Was Redirected", "Indexable", "Title", "Title Length",
        "Meta Description", "Meta Length", "H1 Count", "Word Count", "Internal Links Out",
        "Incoming Internal Links", "Images Total", "Images Missing Alt", "Images Missing Dimensions",
        "Canonical", "Schema Present", "Crawl Depth", "Orphan", "Issues",
    ]
    rows = []
    for p in pages:
        if p.get("status_code") != 200:
            rows.append([p["url"], p.get("status_code") or "unreachable", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", p.get("depth", ""), "", "fetch failed"])
            continue
        issues = compute_page_issues(p, broken_source_pages, protocol_dup_urls)
        rows.append(
            [
                p["url"], p.get("status_code"), p.get("final_url", p["url"]), "Yes" if p.get("was_redirected") else "No",
                "Yes" if p.get("indexable", True) else "No", p.get("title") or "", p.get("title_length", 0),
                p.get("meta_description") or "", p.get("meta_description_length", 0), p.get("h1_count", 0),
                p.get("word_count", 0), p.get("internal_link_count", 0), p.get("incoming_internal_links", 0),
                p.get("images_total", 0), p.get("images_missing_alt", 0), p.get("images_missing_dimensions", 0),
                p.get("canonical") or "", "Yes" if p.get("schema_present") else "No", p.get("depth", 0),
                "Yes" if p.get("is_orphan") else "No", ", ".join(issues) or "none",
            ]
        )
    _write_sheet(ws, headers, rows)
    path = os.path.join(XLSX_DIR, f"page-audit-{_safe_name(site_data['url'])}.xlsx")
    wb.save(path)
    return path


def build_redirects_xlsx(site_data: dict) -> str:
    redirects = site_data.get("redirects_report", [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Redirects"
    headers = ["Original URL", "Final URL", "Status Code", "Hop Count"]
    rows = [[r["original_url"], r["final_url"], r.get("status_code"), r["hop_count"]] for r in redirects]
    _write_sheet(ws, headers, rows)
    path = os.path.join(XLSX_DIR, f"redirects-{_safe_name(site_data['url'])}.xlsx")
    wb.save(path)
    return path


def build_metadata_xlsx(site_data: dict) -> str:
    pages = [p for p in site_data["crawl"]["pages"] if p.get("status_code") == 200]
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"
    headers = ["URL", "Title", "Title Length", "Meta Description", "Meta Length", "Canonical"]
    rows = [
        [p["url"], p.get("title") or "", p.get("title_length", 0), p.get("meta_description") or "", p.get("meta_description_length", 0), p.get("canonical") or ""]
        for p in pages
    ]
    _write_sheet(ws, headers, rows)
    path = os.path.join(XLSX_DIR, f"metadata-{_safe_name(site_data['url'])}.xlsx")
    wb.save(path)
    return path


def build_xlsx_package(site_data: dict) -> dict:
    return {
        "broken_links": build_broken_links_xlsx(site_data),
        "image_alt": build_image_alt_xlsx(site_data),
        "page_audit": build_page_audit_xlsx(site_data),
        "redirects": build_redirects_xlsx(site_data),
        "metadata": build_metadata_xlsx(site_data),
    }
